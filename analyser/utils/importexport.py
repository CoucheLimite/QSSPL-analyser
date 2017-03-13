
import numpy as np
import os
import sys
from shutil import copyfile
import json
import openpyxl as pyxl
import scipy.constants as C
import ruamel.yaml as yaml


def yaml_to_dic(fname):
    '''
    Given a file written in Yaml, returns a dictionary of its contence.

    '''
    with open(os.path.join(fname), 'r') as f:
        bag_of_dics = yaml.safe_load(f.read())

    massive_dic = {}
    for dic in bag_of_dics.values():
        if type(dic) == dict:
            massive_dic.update(dic)

    # get rid of Nones in the massive_dic
    dic = {k: v for k, v in massive_dic.items() if v is not None}

    return dic


class loader_base():

    defaultInputs = {
        'sample': {'doping': 1,
                   'thickness': 1,
                   'reflection': 0.0,
                   'Ai': 1,
                   'temp': 300
                   },
        'measurement': {
            'Fs': 1,
            'Quad': 0.0004338,
            'Lin': 0.03611,
            'Const': 0,
            'waveform': 'blank',
            'gain_pl': 1,
            'gain_gen': 1,
        },
        'analysis': {
            'binning': 1,
            'cropStart': 5,
            'cropEnd': 95,
            'generation': 'generalised',
        }
    }

    inf_ext = '.info'
    old_inf_ext = None
    data_ext = None

    def __init__(self, fname):
        self.directory = os.path.dirname(fname)
        self.rawDataFile = os.path.basename(fname)
        self.inf_fname = self.get_inf_name()
        self.old_inf_fname = self.get_old_inf_name()

    def load_raw_data(self):
        pass

    def load_information(self):
        '''
        Reads the information file form a yaml format.
        This is the default format to be used.
        The version of this below should be just conversation wrapped around this function.
        '''

        # get the information stored in the dictionary
        dic = yaml_to_dic(os.path.join(self.directory, self.inf_fname))

        # get the values required, and make sure they are here
        temp = dict(self.defaultInputs)
        temp.update(dic)

        # return the joined dictionary
        return temp

    def load_processed_data(self):
        print('Still under construction')
        pass

    def write_inf_to_file(self, metadata):
        '''
        Writes the information to a yaml file.
        '''

        # yaml has a problem saving numpy numbers
        # this gets around that by assigning them as a number
        _settings_dic = dict(metadata)

        def ensure_numpy2float(value):
            if isinstance(value, (np.ndarray, np.generic)):
                _settings_dic[dic][key] = float(value)

        # go into the dictionary, check for numpy values, and convert to float
        for dic in metadata:
            for key, value in metadata[dic].items():
                if isinstance(value, (np.ndarray, np.generic)):
                    _settings_dic[dic][key] = float(value)

        # open the file and write it.
        fname = os.path.join(self.directory, self.inf_fname)

        print(_settings_dic)
        with open(fname, 'w') as fname:
            yaml.dump(_settings_dic, fname, default_flow_style=False,
                      indent=4, Dumper=yaml.RoundTripDumper)

    def get_inf_name(self):
        if self.rawDataFile.count(self.data_ext) == 1:
            inf_fname = self.rawDataFile.replace(
                self.data_ext, self.inf_ext)
        else:
            print('Don\'t change the files names')

        return inf_fname

    def get_old_inf_name(self):
        if self.rawDataFile.count(self.data_ext) == 1:
            old_inf_fname = self.rawDataFile.replace(
                self.data_ext, self.old_inf_ext)
        else:
            print('Don\'t change the files names')

        return old_inf_fname


class Load_sinton(loader_base):

    data_ext = '.xlsm'
    old_inf_ext = '.xlsm'

    def load_raw_data(self):
        '''
        Loads a Sinton excel and passes it into a lifetime class, with the
        attributes automatically filled. You still need to check that the Sinton
        excel values were correctly choosen.
        '''

        # define the lifetime class
        # get the measurement data
        file_path = os.path.join(self.directory, self.rawDataFile)

        wb = pyxl.load_workbook(file_path, read_only=True, data_only=True)
        data = self._openpyxl_Sinton2014_ExtractRawDatadata(wb)
        inf = self._openpylx_sinton2014_extractsserdata(wb)

        data.dtype.names = ('time', 'PC', 'gen', 'PL')
        data['PC'] += inf['dark_voltage']

        return data

    def _openpyxl_Sinton2014_ExtractRawDatadata(self, wb):
        '''
            reads the raw data a sinton WCT-120 spreadsheet form the
            provided instance of the openpylx workbook.
        '''

        # make sure the sheet is in the book
        assert 'Calc' in wb.get_sheet_names()

        # get the worksheet
        ws = wb.get_sheet_by_name('Calc')

        # get first section of data
        values1 = np.array([[i.value for i in j] for j in ws['A9':'C133']],
                           dtype=np.float64)

        # add in values so that the background correction works
        for repeat in range(int(values1.shape[0] * 0.1)):
            values1 = np.vstack((values1, values1[-1, :]))
            values1[2, -1] = 0

        headers1 = tuple(
            [[j.value for j in i] for i in ws['A8':'C8']][0])

        # get second section of data
        values2 = np.zeros(values1.shape[0])
        headers2 = ('PL',)

        # form into one array with names
        values = np.vstack((values1.T, values2)).T
        headers = headers1 + headers2

        dtype = [('a', np.float64), ('b', np.float64),
                 ('c', np.float64), ('d', np.float64)]

        Out = values.copy().view(dtype=dtype).reshape(values.shape[0],)

        time_diff = Out['a'][2] - Out['a'][1]
        for i in range(Out['a'].shape[0]):
            Out['a'][i] = i * time_diff - time_diff

        return Out

    def _openpylx_sinton2014_extractsserdata(self, wb):
        '''
        reads the measurement and sample information from a sinton WCT-120
        spreadsheet form the provided instance of the openpylx workbook.
        '''
        # make sure the sheet is in the book
        # get the worksheet
        assert 'User' in wb.get_sheet_names()
        assert 'Settings' in wb.get_sheet_names()

        ws = wb.get_sheet_by_name('User')

        # Grabbing the data and assigning it a nae

        user_set = {
            'thickness': float(ws['B6'].value),
            'doping': float(ws['J9'].value),
            'sample_type': ws['D6'].value.encode('utf8'),
            'optical_constant': float(ws['E6'].value),
        }

        user_set['reflection'] = (1 - user_set['optical_constant']) * 100

        # makes a reference to the RawData page
        ws = wb.get_sheet_by_name('Settings')

        # grabs the Cell ref and terns it into a FS
        sys_set = {
            'Fs': 0.038 / C.e / float(ws['C5'].value),
        }

        # make one dic
        user_set.update(sys_set)

        ws = wb.get_sheet_by_name('Calc')
        sys_set = {
            'dark_voltage': float(ws['B166'].value),
        }

        user_set.update(sys_set)

        return user_set

    def load_information(self):
        '''
        This creates an extra inf file, which is used to stored the
        information used to analyse the file. If the file exists, the
        information is pulled from that file.
        '''

        # if the yaml file does not exists the default load
        if not os.path.isfile(os.path.join(self.directory, self.inf_fname)):

            file_path = os.path.join(self.directory, self.old_inf_fname)
            wb = pyxl.load_workbook(file_path, read_only=True, data_only=True)

            inf_dic = self._openpylx_sinton2014_extractsserdata(wb)
            temp = dict(self.defaultInputs)
            temp.update(inf_dic)

            # write to the yaml file.
            super().write_inf_to_file(temp)

        # read the yaml file
        return super().load_information()

    def load_processed_data(self):
        DataFile = self.DataFile[:-13] + '.dat'
        return np.genfromtxt(self.directory + DataFile, usecols=(0, 1, 8, 9), unpack=True, delimiter='\t', names=('Deltan_PC', 'Tau_PC', 'Deltan_PL', 'Tau_PL'))


class Load_QSSPL_File_LabView(loader_base):

    old_inf_ext = '.inf'
    data_ext = '_Raw Data.dat'

    def load_raw_data(self):
        fname = os.path.join(self.directory, self.rawDataFile)
        return np.genfromtxt(fname,
                             names=('time', 'PC', 'gen', 'PL'))

    def load_information(self):
        '''
        Just a convert to put things into the new yaml format.
        '''

        if not os.path.isfile(os.path.join(self.directory, self.inf_fname)):

            file_path = os.path.join(self.directory, self.old_inf_fname)

            Cycles, dump, Frequency, LED_Voltage, dump, dump, dump, dump, DataPoints, dump = np.genfromtxt(
                file_path, skip_header=20, skip_footer=22, delimiter=':', usecols=(1), autostrip=True, unpack=True)
            Waveform, LED_intensity = np.genfromtxt(
                file_path, skip_header=31, skip_footer=20, delimiter=':', usecols=(1), dtype=None, autostrip=True, unpack=True)

            l = np.genfromtxt(
                file_path, skip_header=36, delimiter=':', usecols=(1))

            dic = {}
            sample = {}
            measurement = {}
            analysis = {}

            measurement['Cycles'] = Cycles
            measurement['Frequency'] = Frequency
            measurement['LED_Voltage'] = LED_Voltage
            measurement['DataPoints'] = DataPoints
            measurement['Waveform'] = Waveform.decode()
            measurement['LED_intensity'] = LED_intensity.decode()
            measurement['Fs'] = l[7]
            measurement['Quad'] = l[13]
            measurement['Lin'] = l[14]

            sample['Doping'] = l[9]
            sample['Ai'] = l[6]
            sample['Thickness'] = l[12]
            sample['sample'] = (1 - l[16]) * 100

            analysis['Binning'] = int(l[2])

            dic['sample'] = sample
            dic['measurement'] = measurement
            dic['analysis'] = analysis

            # write to the yaml file.
            super().write_inf_to_file(dic)

        return super().load_information()

    def load_processed_data(self):
        DataFile = self.DataFile[:-13] + '.dat'
        return np.genfromtxt(self.directory + DataFile, usecols=(0, 1, 8, 9), unpack=True, delimiter='\t', names=('Deltan_PC', 'Tau_PC', 'Deltan_PL', 'Tau_PL'))


class Load_QSSPL_File_Python(loader_base):

    old_inf_ext = '.inf'
    data_ext = '.Raw Data.dat'

    def load_raw_data(self):
        fname = os.path.join(self.directory, self.rawDataFile)
        data = np.genfromtxt(fname, unpack=True, names=True, delimiter='\t')

        s = np.array([])
        dic = {'Time_s': 'time', 'Gen_V': 'gen', 'Generation_V': 'gen',
               'PL_V': 'PL', 'PC_V': 'PC'}

        for i in np.array(data.dtype.names):

            s = np.append(s, dic[i])

        data.dtype.names = s

        return data

    def num(self, s):
        try:
            return float(s)
        except ValueError:
            return s

    def load_information(self):
        '''
        Converts the saved inf file into the new format
        '''

        if not os.path.isfile(os.path.join(self.directory, self.inf_fname)):
            file_path = os.path.join(self.directory, self.old_inf_fname)
            with open(file_path, 'r') as f:
                s = f.read()

            # removes double line endings
            s = s.replace('\n\n', '\n')
            # removes tabs
            s = s.replace('\t', '    ')
            # removes the first to header lines
            s = '\n'.join(s.split('\n')[2:])
            measurement = yaml.safe_load(s)
            dic = {
                'measurement': measurement
            }
            super().write_inf_to_file(dic)

        return super().load_information()


# class Python_auto_load():
#
#     def __init__(self, fname):
#         self.directory = os.path.dirname(fname)
#         self.rawDataFile = os.path.basename(fname)
#
#     def load_raw_data(self):
#         '''
#         Loads the measured data from the data file.
#         This has the file extension tsv (tab seperated values)
#
#         from a provided file name,
#         takes data and outputs data with specific column headers
#         '''
#
#         # get data, something stange was happening with os.path.join
#         file_location = os.path.normpath(
#             os.path.join(self.directory, self.rawDataFile))
#
#         data = np.genfromtxt(
#             os.path.join(file_location),
#             unpack=True, names=True, delimiter='\t')
#
#         # string to convert file names to program names
#         dic = {'Time_s': 'time', 'Generation_V': 'gen',
#                'PL_V': 'PL', 'PC_V': 'PC'}
#
#         # create empty array
#         s = np.array([])
#
#         # build array of names, in correct order
#         for i in np.array(data.dtype.names):
#             s = np.append(s, dic[i])
#
#         # assign names
#         data.dtype.names = s
#
#         return data
#
#     def num(self, s):
#         '''
#         converts s to a number, or returns s
#         '''
#         try:
#             return float(s)
#         except ValueError:
#             return s
#
#     def load_information(self):
#         # print 'Still under construction'
#
#         # replace the ending with a new ending
#         self.inf_fname = self.get_inf_name()
#
#         # These are adjustment Values, required by the following
#         temp_dic = {'doping': None,
#                     'thickness': None,
#                     'binning_pp': 1,
#                     'reflection': None,
#                     'doping_type': None,
#                     'Fs': 1.0727E+20,
#                     'Ai': 3e22,
#                     'Quad': 0.0004338,
#                     'Lin': 0.03611,
#                     'Const': 0,
#                     'cropStart': None,
#                     'cropEnd': None,
#                     'waveform': None,
#                     'temp': 300,
#                     'gain_pl': 1,
#                     'gain_gen': 1,
#                     }
#
#         with open(os.path.join(self.directory, self.inf_fname), 'r') as f:
#             bag_of_dics = yaml.load(f.read())
#
#         massive_dic = {}
#         for dic in bag_of_dics.values():
#             massive_dic.update(dic)
#
#         # get rid of Nones in the massive_dic
#         dic = {k: v for k, v in massive_dic.items() if v is not None}
#
#         temp_dic.update(dic)
#
#         return temp_dic
#


# class TempDep_loads():
#
#     def __init__(self, fname):
#         self.directory = os.path.dirname(fname)
#         self.rawDataFile = os.path.basename(fname)
#
#     def load_raw_data(self):
#         '''
#         Loads the measured data from the data file.
#         This has the file extension tsv (tab seperated values)
#
#         from a provided file name,
#         takes data and outputs data with specific column headers
#         '''
#
#         # get data, something stange was happening with os.path.join
#         file_location = os.path.normpath(
#             os.path.join(self.directory, self.rawDataFile))
#
#         data = np.genfromtxt(
#             os.path.join(file_location),
#             unpack=True, names=True, delimiter='\t')
#
#         # string to convert file names to program names
#         dic = {'Time_s': 'time', 'Generation_V': 'gen',
#                'PL_V': 'PL', 'PC_V': 'PC'}
#
#         # create empty array
#         s = np.array([])
#
#         # build array of names, in correct order
#         for i in np.array(data.dtype.names):
#             s = np.append(s, dic[i])
#
#         # assign names
#         data.dtype.names = s
#
#         return data
#
#     def num(self, s):
#         '''
#         converts s to a number, or returns s
#         '''
#         try:
#             return float(s)
#         except ValueError:
#             return s
#
#     def load_information(self):
#         # print 'Still under construction'
#
#         # replace the ending with a new ending
#         self.inf_fname = self.get_inf_name()
#
#         # These are adjustment Values, requried by the following
#         temp_list = {'doping': 1,
#                      'thickness': 1,
#                      'binning_pp': 1,
#                      'reflection': 0.0,
#                      'Fs': 1,
#                      'Ai': 1,
#                      'quad': 0.0004338,
#                      'lin': 0.03611,
#                      'constant': 0,
#                      'cropStart': None,
#                      'cropEnd': None,
#                      'waveform': None,
#                      }
#
#         with open(os.path.join(self.directory, self.inf_fname), 'r') as f:
#             file_contents = f.read()
#             List = json.loads(file_contents)
#
#         List.update(temp_list)
#
#         return List


class LoadData():

    fname = ''
    File_Type = ''

    loaders = [Load_sinton,
               Load_QSSPL_File_LabView,
               Load_QSSPL_File_Python]

    def __init__(self, fname):
        self.fname = fname

    def _obtain_loader(self):

        for loader in self.loaders:
            if loader.data_ext in self.fname:
                break

        return loader(self.fname)

    def load(self):
        '''
        Returns a data file containing the measured data and an dictionary
        containing measurement information.
        '''
        loader = self._obtain_loader()
        return loader.load_raw_data(), loader.load_information()

    def load_processed_data(self):
        loader = self._obtain_loader()
        return loader.load_processed_data()

    def save(self, data, settings):
        '''
        Saves the new analysis information into the inf file.
        '''
        LoadClass = self.obtain_loader()
        return LoadClass.write_inf_to_file(settings)
